import openpyxl
import pprint
# import qrcode

wb = openpyxl.load_workbook('file_data_Copy.xlsx')
wrkb = openpyxl.Workbook()

sheet = wb['Sheet1']
ws = wrkb.worksheets[0]

max = sheet.max_row
print(max)

for i in range(max):
    a = 'B'+str(i+1)
    #print(a)
    cell = sheet[a]
    #print(cell.value)
    # img = qrcode.make('W' + cell.value)
    name = "W" + cell.value
    # img.save(f'CreateQR/{name}.png')


    # c = 'C' + str(i+1)
    # wb.insert_image(c, 'CreateQR/MyQRCode{}.png'.format(i+1), {'x_scale': 0.1, 'y_scale': 0.1})

# img = openpyxl.drawing.image.Image('MyQRCode1.png')
# img.anchor = 'C1'
# ws.add_image(img)
wb.save('file_data_Copy.xlsx')